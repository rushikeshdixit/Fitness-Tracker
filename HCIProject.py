from flask import Flask, render_template, request, redirect, jsonify
from flask_mysqldb import MySQL
from sklearn import svm
from sklearn.neighbors import KNeighborsClassifier
import numpy as np
import random
import pandas as pd
import xlrd
import sklearn
import openpyxl
from openpyxl import load_workbook
import os
from openpyxl import load_workbook
from sklearn import linear_model
import matplotlib.pyplot as plt
import numpy as np
import xlwt
from scipy import sparse
# from xlutils.copy import copy


app = Flask(__name__)


# MySQL configurations

app.config['MYSQL_USER'] = '<enter_username>'
app.config['MYSQL_PASSWORD'] = '<enter_password>'
app.config['MYSQL_DB'] = '<enter_db>'
app.config['MYSQL_HOST'] = '<enter_hostname>'
mysql = MySQL(app)
print "success"
age=""
age1 = {"AGE":"1"}

def get_user_credentials(username):
    conn = mysql.connection
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usercred WHERE username = '"+username+"';")
    data = cursor.fetchall()
    return data

def insert_user_credentials(username, password):
    conn = mysql.connection
    cursor = conn.cursor()
    cursor.execute("insert into usercred values('"+username+"', '"+password+"');")
    conn.commit()
    print ('Added the user to the table.')

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'GET':
       return render_template('index.html')
    if request.method == 'POST':
        username = request.form['username']
        print("username is: " + username)
        storedusername = [i[0] for i in get_user_credentials(username)]
        print("stored username is:", storedusername)
        password = request.form['password']
        print("password is:" + password)
        storedpassword = [i[1] for i in get_user_credentials(username)]
        print("storedpassword is:", storedpassword)
        if username in storedusername and password in storedpassword:
            return redirect('mainPage')
        return "<h1>Account not present</h1>"

@app.route('/signUp', methods=['POST'])
def signUp():
    # read the posted values from the UI
    username = request.form['username']
    print("username is: "+username)
    storedusername=[i[0] for i in get_user_credentials(username)]
    print("storedusername is:", storedusername)
    password = request.form['password']
    print password
    if username not in storedusername:
        insert_user_credentials(username, password)
        return "<h1>user inserted</h1>"
    return "<h1>username already present</h1>"

@app.route('/mainPage', methods=['GET', 'POST'])
def mainPage():
        if request.method == 'GET':
            print 'get method executed'
            # slider = request.json['waist']
            # print 'Slider is', slider
            return render_template('main.html')
        if request.method == 'POST':
            obj1=request.json['waist']
            print "waist is", obj1
            obj2 = request.json['height']
            print "height is", obj2
            obj3 = request.json['age']
            print "age is", obj3
            ratio=float(obj1)/float(obj2)
            print ratio
            age1["AGE"]=obj3
            # jsonify(''ratio)
            # dict={'waist':obj1, 'height':obj2, 'age':obj3, 'ratio':ratio}
           # floatRatio=float(ratio)
            #print floatRatio
            return jsonify(ratio)

@app.route('/appleExercises1', methods=['GET', 'POST'])
def appleExercises1():
        if request.method == 'GET':
            return render_template('appleExercises1.html')
        if request.method == 'POST':
            age=request.json['age']
            print age
            return render_template('appleExercises1.html')
            
@app.route('/appleExercises2', methods=['GET', 'POST'])
def appleExercises2():
        if request.method == 'GET':
            return render_template('appleExercises2.html')
        if request.method == 'POST':
            age=request.json['age']
            print age
            return render_template('appleExercises2.html')

@app.route('/weight_birdDog', methods=['GET', 'POST'])
def weight_birdDog():
        if request.method == 'GET':
            return render_template('weight_birdDog.html')
        if request.method == 'POST':
            age=request.json['age']
            print age
            return render_template('weight_birdDog.html')

@app.route('/rear_lounge', methods=['GET', 'POST'])
def rear_lounge():
        if request.method == 'GET':
            return render_template('rear_lounge.html')
        if request.method == 'POST':
            age=request.json['age']
            print age
            return render_template('rear_lounge.html')

@app.route('/pearExercises', methods=['GET', 'POST'])
def pearExercises():
    if request.method == 'GET':
        return render_template('pearExercises.html')
    if request.method == 'POST':
        age = request.json['age']
        print age
        return render_template('pearExercises.html')
        
@app.route('/pearExercises2', methods=['GET', 'POST'])
def pearExercises2():
    if request.method == 'GET':
        return render_template('pearExercises2.html')
    if request.method == 'POST':
        age = request.json['age']
        print age
        return render_template('pearExercises2.html')
        
@app.route('/pear_plankleg', methods=['GET', 'POST'])
def pear_plankleg():
        if request.method == 'GET':
            return render_template('pear_plankleg.html')
        if request.method == 'POST':
            age=request.json['age']
            print age
            return render_template('pear_plankleg.html')

@app.route('/pear_vpull', methods=['GET', 'POST'])
def pear_vpull():
        if request.method == 'GET':
            return render_template('pear_vpull.html')
        if request.method == 'POST':
            age=request.json['age']
            print age
            return render_template('pear_vpull.html')

@app.route('/user_input', methods=['GET', 'POST'])
def user_input():
    if request.method == 'GET':
        return render_template('user_input.html')
    if request.method == 'POST':
        return render_template('user_input.html')


#Machine Learning code to predict diabetes

# @app.route('/', methods=['GET'])
# def index():
#     if(request.method == 'GET'):
#         return render_template('user_input.html')
#
#
@app.route('/getdata', methods=['GET', 'POST'])
def check():
    print("I came here")
    preg = request.form.get('pregnancy')
    bloodP =request.form.get('BloodPressure')
    skinT = request.form.get('skinthickness')
    insulin = request.form.get('insulin')
    bmi = request.form.get('bmi')
    age = request.form.get('age')

    print("Preg = ", preg, " Blood Pressure= ", bloodP)
    print("==============================")

    appendDataToExcel(preg,bloodP,skinT, insulin, bmi, age)
    print("=========Text Excel Saved=======")
    print('Data in excel is : ')

    var = precitMe()
    print("var is",var)
    msg = ''
    if var < 150 and var >= 90:
        msg = "Your sugar level is normal! Keep up the good work"
    elif var >= 160 and var <= 240:
        msg = "Your blood sugar is too high. Work on lowering down your sugar intake"
    elif var >= 240 and var <= 300:
        msg = "Your sugar level is out of control"
    else:
        msg = "Your life is in danger"


    return render_template('display_data.html', msg = msg)

def appendDataToExcel(preg,bloodP,skinT, insulin, bmi, age):
    print("Entered Append Function")
    file = '/Users/Rushi/PycharmProjects/HCIProject/static/diabetes_data.xlsx'
    new_row = [preg,bloodP,skinT, insulin,bmi, age]
    print new_row
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.get_sheet_by_name('diabetes')
    row = ws.max_row +1
    for col, entry in enumerate(new_row, start=1):
        print('col is', col, 'entry is', entry)
        ws.cell(row=row, column=col, value=entry)
    wb.save(file)
    wb.close()
    print("Leaving Append Function")



def precitMe():
    print('Into the function predictMe.... Started Function')
    value = predictDiabetes()
    var = int(value[0])
    wb = load_workbook(filename='/Users/Rushi/PycharmProjects/HCIProject/static/diabetes_data.xlsx')
    ws = wb.worksheets[0]
    row = ws.max_row
    ws.cell(row=row, column=7).value = var
    wb.close()
    print("====================================================")
    print(var)

    return var

def predictDiabetes():
    print("lol i came here")
    # database = xlrd.open_workbook('/Users/Rushi/Downloads/Diabetes-Prediction-for-Women-master/diabetes_data.xlsx')
    # wb = load_workbook(filename='/Users/Rushi/Downloads/Diabetes-Prediction-for-Women-master/diabetes_data.xlsx')
    # ws= wb.worksheets[0]
    # print 'Database is :', ws
    col_names = ['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age',
                 'Glucose']
    print(col_names)
    pima = pd.read_excel("/Users/Rushi/PycharmProjects/HCIProject/static/diabetes_data.xlsx", names=col_names, engine='xlrd')
    print 'pima is', pima
    per = 0.55
    plot = pd.DataFrame(col_names)
    print 'plot is :', plot
    train = pima[:((int)(len(plot) * per))]
    xtrain = train[
        ['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age']]
    ytrain = train[['Glucose']]

    test = pima[((int)(len(plot) * per)):]
    print 'Test is', test
    xtest = test[['Pregnancies', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI', 'Age']]
    print('last value ofXtest is ', xtest)
    ytest = test[['Glucose']]
    dataframeHeading = plot.head()

    ols = sklearn.linear_model.LinearRegression()
    model = ols.fit(xtrain, ytrain)

    predict = model.predict(xtest)
    print("==Value of model predict===")

    length = len(predict) -1
    print(predict[length])
    print("This is the value?")
    print("Ends Here")
    pima.update(pima)

    return predict[length]


@app.route('/foodItem', methods=['GET', 'POST'])
def foodItem():
    print("entered food Item")
    newValue = request.form.get('foodValue')
    print(newValue)


    file = '/Users/Rushi/PycharmProjects/HCIProject/static/Foods.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.get_sheet_by_name('Sheet1')


    sugar = 0

    print("11111")
    for i in range(1, ws.max_row):

        if ws.cell(row=i, column=1).value == newValue:
            print("entered this thing")
            print(sugar)
            sugar = ws.cell(row=i, column=3).value
    sugar = 2500-sugar
    print("22222222222")
    msg = ''
    if sugar < 270 and sugar >= 135:
        msg = "No sugar for you! You are done for today"
    elif sugar >= 135 and sugar <= 95:
        msg = "You can have some Tofu, beans, oats"
    elif sugar >= 95 and sugar <= 50:
        msg = "You can have fruits"
    else:
        msg = "You can eat anything upto 2500 mg of sugar! "

    print("left food item")

    return render_template('final.html', msg = msg, sugar = sugar)






if __name__ == '__main__':
    app.run()
